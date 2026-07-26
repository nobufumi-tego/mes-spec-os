"""cache.py の単体テスト

ENAA 実ファイルなしで実行できるよう、G25-029-2 と同じレイアウトの
ダミー Excel を一時ディレクトリに生成してテストする。

実行方法（リポジトリのルートで）:

    PYTHONIOENCODING=utf-8 uv run --with pytest --with openpyxl --no-project \
        python -m pytest tests/ -q
"""
import json
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'usecases', 'MES仕様書'))
import cache as cache_mod

SHEET_NAME = 'ENAA MES標準業務機能リスト (最新）'

# (id, Division, BC, SubBC, BP, MES対象, 業務オペレーション, 補足,
#  代替可能システム, ERP, PLM, スケジューラー, コントロールシステム, その他システム)
DUMMY_ROWS = [
    ('A-10-10-01', '生産管理', '生産計画立案', '大日程計画', '計画作成', '○',
     '月次の生産計画を立案する', None, 'SCP', 'I', None, 'O', None, None),
    ('A-10-10-02', '生産管理', '生産計画立案', '小日程計画', '計画調整', '△',
     '日々の計画を調整する', '補足あり', 'スケジューラー', None, None, 'I/O', None, None),
    ('B-20-10-01', '製造', '製造実行', '作業実施', '実績登録', '〇',
     '製造実績を登録する', None, None, 'O', None, None, 'I', None),
    ('B-20-10-02', '製造', '製造実行', '作業実施', '帳票出力', '×',
     '紙帳票を出力する', None, 'BI', None, None, None, None, 'O'),
    ('C-30-10-01', '品質管理', '検査', '受入検査', '検査記録', '※',
     '検査結果を記録する', '品質 Division 参照', 'QMS', None, None, None, None, None),
]


def _build_dummy_xlsx(path, with_custom=False):
    """G25-029-2 と同じレイアウト（B2 タイトル / 5 行目ヘッダ / 6 行目からデータ）のダミーを作る"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.cell(2, 2, 'ダミー標準業務一覧 2025 年 9 月 30 日 版')
    headers = ['id', 'Division', 'BC', 'SubBC', 'BP', 'MES対象', '業務オペレーション',
               '補足', '代替可能システム', 'ERP', 'PLM', 'スケジューラー',
               'コントロールシステム', 'その他システム']
    for i, h in enumerate(headers):
        ws.cell(5, 2 + i, h)
    if with_custom:
        ws.cell(5, 16, '担当部門')
    for r, row in enumerate(DUMMY_ROWS, start=6):
        for c, v in enumerate(row, start=2):
            ws.cell(r, c, v)
    if with_custom:
        ws.cell(6, 16, '生産管理部')
    wb.save(path)


@pytest.fixture
def dummy_env(tmp_path, monkeypatch):
    """ダミー Excel と一時キャッシュパスへ cache.py の参照先を差し替える"""
    xlsx = tmp_path / 'mes_g25-029-2.xlsx'
    _build_dummy_xlsx(str(xlsx))
    monkeypatch.setattr(cache_mod, 'ENAA_XLSX', str(xlsx))
    monkeypatch.setattr(cache_mod, 'CACHE_FILE', str(tmp_path / '.cache' / 'enaa_full.json'))
    return xlsx


def test_build_rows_and_version(dummy_env):
    c = cache_mod.get_cache()
    assert c['_meta']['row_count'] == len(DUMMY_ROWS)
    assert c['_meta']['version'] == '2025-09-30'
    assert c['rows'][0]['id'] == 'A-10-10-01'
    assert c['rows'][0]['Division'] == '生産管理'


def test_cache_hit_returns_same_snapshot(dummy_env):
    first = cache_mod.get_cache()
    second = cache_mod.get_cache()
    assert second['_meta']['cached_at'] == first['_meta']['cached_at']


def test_mtime_change_triggers_rebuild(dummy_env):
    first = cache_mod.get_cache()
    os.utime(str(dummy_env), (0, 0))  # mtime を過去に変更
    second = cache_mod.get_cache()
    assert second['_meta']['excel_mtime'] != first['_meta']['excel_mtime']


def test_corrupted_cache_rebuilds_silently(dummy_env):
    cache_mod.get_cache()
    with open(cache_mod.CACHE_FILE, 'w', encoding='utf-8') as f:
        f.write('{ broken json')
    c = cache_mod.get_cache()
    assert c['_meta']['row_count'] == len(DUMMY_ROWS)


def test_missing_excel_raises(tmp_path, monkeypatch):
    monkeypatch.setattr(cache_mod, 'ENAA_XLSX', str(tmp_path / 'nonexistent.xlsx'))
    monkeypatch.setattr(cache_mod, 'CACHE_FILE', str(tmp_path / '.cache' / 'enaa_full.json'))
    with pytest.raises(FileNotFoundError):
        cache_mod.get_cache()


def test_filters(dummy_env):
    c = cache_mod.get_cache()
    assert len(cache_mod.filter_by_division(c, '生産管理')) == 2
    assert len(cache_mod.filter_by_subbc(c, '作業実施')) == 2
    assert cache_mod.get_by_id(c, 'C-30-10-01')['BC'] == '検査'
    assert cache_mod.get_by_id(c, 'Z-99-99-99') is None


def test_custom_columns(dummy_env, tmp_path, monkeypatch):
    c = cache_mod.get_cache()
    assert not cache_mod.has_custom_data(c)

    xlsx2 = tmp_path / 'custom.xlsx'
    _build_dummy_xlsx(str(xlsx2), with_custom=True)
    monkeypatch.setattr(cache_mod, 'ENAA_XLSX', str(xlsx2))
    monkeypatch.setattr(cache_mod, 'CACHE_FILE', str(tmp_path / '.cache2' / 'enaa_full.json'))
    c2 = cache_mod.get_cache()
    assert cache_mod.has_custom_data(c2)
    rows = cache_mod.get_custom_rows(c2)
    assert len(rows) == 1
    assert rows[0]['_custom']['担当部門'] == '生産管理部'


def test_classify_mes_target():
    assert cache_mod.classify_mes_target('○') == 'on'
    assert cache_mod.classify_mes_target('〇') == 'on'   # 異体字
    assert cache_mod.classify_mes_target('△') == 'partial'
    assert cache_mod.classify_mes_target('×') == 'off'
    assert cache_mod.classify_mes_target('※') == 'reference'
    assert cache_mod.classify_mes_target(None) == 'unknown'
    assert cache_mod.classify_mes_target(' ○ ') == 'on'  # 前後空白


def test_count_mes_targets(dummy_env):
    c = cache_mod.get_cache()
    counts = cache_mod.count_mes_targets(c['rows'])
    assert counts == {'on': 2, 'partial': 1, 'off': 1, 'reference': 1,
                      'unknown': 0, 'total': 5}
