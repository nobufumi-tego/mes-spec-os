"""ENAA Excel データのキャッシュヘルパー（v0.2）

ENAA Excel (mes_g25-029-2.xlsx) を 1 度読み込んで `outputs/.cache/enaa_full.json`
にキャッシュし、以降のレシピ呼び出しはキャッシュから読む。Excel の mtime が
変わった場合（利用者が ENAA から再ダウンロードした等）は自動で再構築。

使い方（レシピ側から）:

    import sys, os, json
    THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, THIS_DIR)
    from cache import get_cache, filter_by_division, filter_by_subbc, get_by_id

    cache = get_cache()              # キャッシュヒット or 自動構築
    rows = cache['rows']             # 全 443 行（dict のリスト）
    meta = cache['_meta']            # バージョン・mtime 等

    # 便利フィルタ
    seizou_rows = filter_by_division(cache, '製造')
    seizou_jisshi = filter_by_subbc(cache, '製造実施')
    target = get_by_id(cache, 'B-30-30-01')

設計:
- パスは self-locating（cache.py の位置からの相対パスで決定）。private 開発時
  と public 利用時の両方で動作する
- Excel が見つからない場合は FileNotFoundError を送出（呼び出し元でフォール
  バック処理）
- キャッシュファイルが破損している場合は黙って再構築
- 出力は ensure_ascii=False のコンパクト JSON（~250KB を想定）
"""
import os
import json
import datetime
import re
from typing import Dict, Any, List, Optional


# === パス解決（self-locating）===
# cache.py はOS_ROOT/usecases/MES仕様書/cache.py に配置されている前提
_THIS_FILE = os.path.abspath(__file__)
_OS_ROOT = os.path.abspath(os.path.join(os.path.dirname(_THIS_FILE), '..', '..'))

ENAA_XLSX = os.path.join(_OS_ROOT, 'reference', 'ENAA', 'mes_g25-029-2.xlsx')
CACHE_FILE = os.path.join(_OS_ROOT, 'outputs', '.cache', 'enaa_full.json')


# === 列マッピング（標準業務一覧 G25-029-2）===
# col 2-15 が ENAA 定義、col 16+ は利用者カスタム領域
COLUMN_MAP = {
    2: 'id',                      # 例: A-10-10-01
    3: 'Division',                # 業務のくくり（12 大分類）
    4: 'BC',                      # Business Category（業務分類）
    5: 'SubBC',                   # Sub Business Category
    6: 'BP',                      # Business Process
    7: 'MES対象',                 # ○ / ×
    8: '業務オペレーション',
    9: '補足',
    10: '代替可能システム',
    11: 'ERP',                    # Integration Sample
    12: 'PLM',
    13: 'スケジューラー',
    14: 'コントロールシステム',
    15: 'その他システム',
}


def _build_cache() -> Dict[str, Any]:
    """ENAA Excel を読み込んで cache 構造を構築"""
    from openpyxl import load_workbook

    wb = load_workbook(ENAA_XLSX, data_only=True)
    ws = wb['ENAA MES標準業務機能リスト (最新）']

    rows: List[Dict[str, Any]] = []
    for r in range(6, ws.max_row + 1):
        div = ws.cell(r, 3).value
        if not div:
            continue  # Division が空の行はスキップ

        row: Dict[str, Any] = {}

        # 標準カラム（col 2-15）
        for col, alias in COLUMN_MAP.items():
            v = ws.cell(r, col).value
            row[alias] = str(v).strip() if v is not None else None

        # 利用者カスタムカラム（col 16+、ENAA Quick Start P13 推奨）
        custom: Dict[str, str] = {}
        for col in range(16, ws.max_column + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            header = ws.cell(5, col).value
            key = str(header).strip() if header else f'col_{col}'
            custom[key] = str(v).strip()
        if custom:
            row['_custom'] = custom

        rows.append(row)

    # バージョン情報を B2 から抽出
    title = str(ws.cell(2, 2).value or '')
    m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', title)
    version = (
        f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
        if m else None
    )

    return {
        '_meta': {
            'excel_path': ENAA_XLSX,
            'excel_mtime': os.path.getmtime(ENAA_XLSX),
            'excel_size': os.path.getsize(ENAA_XLSX),
            'cached_at': datetime.datetime.now().isoformat(timespec='seconds'),
            'version': version,
            'row_count': len(rows),
            'cache_schema_version': '1',
        },
        'rows': rows,
    }


def get_cache(force_rebuild: bool = False) -> Dict[str, Any]:
    """ENAA キャッシュを取得（必要なら自動構築）

    Args:
        force_rebuild: True にすると mtime 判定を無視して必ず再構築

    Returns:
        cache dict with '_meta' and 'rows' keys

    Raises:
        FileNotFoundError: ENAA Excel が見つからない場合
    """
    if not os.path.exists(ENAA_XLSX):
        raise FileNotFoundError(
            f'ENAA Excel not found: {ENAA_XLSX}\n'
            f'reference/ENAA/README.md の手順に従って配置してください'
        )

    # キャッシュヒット判定
    if not force_rebuild and os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            cached_mtime = cache.get('_meta', {}).get('excel_mtime')
            current_mtime = os.path.getmtime(ENAA_XLSX)
            if cached_mtime == current_mtime:
                return cache  # ヒット
        except (json.JSONDecodeError, KeyError, OSError):
            pass  # 破損キャッシュは黙って再構築

    # キャッシュミス → 構築
    cache = _build_cache()
    os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, separators=(',', ':'))
    return cache


# === 便利フィルタ ===

def filter_by_division(cache: Dict[str, Any], division: str) -> List[Dict[str, Any]]:
    """指定 Division に属する全行を返す"""
    return [r for r in cache['rows'] if r.get('Division') == division]


def filter_by_subbc(cache: Dict[str, Any], subbc: str) -> List[Dict[str, Any]]:
    """指定 SubBC に属する全行を返す（同名 SubBC が複数 BC にあっても全件返す）"""
    return [r for r in cache['rows'] if r.get('SubBC') == subbc]


def get_by_id(cache: Dict[str, Any], bp_id: str) -> Optional[Dict[str, Any]]:
    """指定 ID（A-10-10-01 形式）の行を返す（見つからなければ None）"""
    for r in cache['rows']:
        if r.get('id') == bp_id:
            return r
    return None


def has_custom_data(cache: Dict[str, Any]) -> bool:
    """P 列以降の利用者カスタム書き込みが存在するか"""
    return any(r.get('_custom') for r in cache['rows'])


def get_custom_rows(cache: Dict[str, Any]) -> List[Dict[str, Any]]:
    """利用者カスタム書き込みのある行のみ返す"""
    return [r for r in cache['rows'] if r.get('_custom')]


if __name__ == '__main__':
    # 動作確認用: python cache.py で実行するとメタ情報を表示
    import sys
    cache = get_cache()
    print(json.dumps(cache['_meta'], ensure_ascii=False, indent=2))
    print(f'Total rows: {len(cache["rows"])}')
    print(f'Has custom data: {has_custom_data(cache)}')
